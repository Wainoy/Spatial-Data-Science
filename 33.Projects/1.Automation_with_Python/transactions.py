#project 1
import openpyxl as xl #using an alias name to enhance readability
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1'] #access a sheet
cell = sheet['a1']
#or
cell = sheet.cell(1,1) #i.e. (row, column) to get the value stored in a given cell
print(cell.value)
print(sheet.max_row) #print number of row in the excel file

for row in range(2, sheet.max_row + 1):
     sheet.cell(row,3)
     corrected_price = cell.value*0.9
     corrected_price_cell =  sheet.cell(row, 4)
     corrected_price_cell.value = corrected_price


wb.save('transactions2.xslx') 

